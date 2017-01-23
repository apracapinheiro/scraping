import requests
import bs4
import time
import openpyxl


NOME_SITE = 'Lamas'

headers = {'user-agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_9_5)'}  # MAC
# headers = {'user-agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/55.0.2883.87 Safari/537.36'} # CHROME
# headers = {'user-agent': 'Mozilla/5.0 (Windows NT 10.0; WOW64; rv:47.0) Gecko/20100101 Firefox/47.0'}  # FIREFOX

wb = openpyxl.Workbook()
sheet = wb.get_active_sheet()
sheet.title = 'Precos maltes ' + NOME_SITE


def request_site(url, headers, num_pagina):
    """
    Realiza a requisicao da pagina
    :param url: url da página que será realizado o scraping
    :param headers: cabeçalhos contendo informacoes sobre o browser. Necessario para passar pela validacao no servidor
    alvo
    :param num_pagina:
    :return:
    """
    if not num_pagina:
        num_pagina = 1
    time.sleep(3)
    res = requests.get(url, headers=headers)
    status_request = res.raise_for_status()
    return res, status_request


def scraping_insumos(requisicao, status, num_pagina, lista_insumos):
    """
    Funcao que realiza o scraping dos insumos.
    :param requisicao: objeto requisicao da pagina
    :param status: stauts da requisicao, só continua se o valor for None. Valores diferentes de None indicam o erro
    que o servidor enviou
    :param num_pagina: numero da pagina que sera realizado o scraping
    :param lista_insumos: lista que contem os insumos das outras paginas que ja passaram por scraping
    :return: lista de insumos com todas as paginas.
    """
    while status == None:
        pagina = bs4.BeautifulSoup(requisicao.text)
        insumo = pagina.select('div[class="list-conteiner-name"]')
        precos = pagina.select('span[class="regular-price"]')
        proximo = pagina.select('a[class="next i-next"]')

        if len(precos) > 0:
            for i in range(len(precos)):
                nome_insumo = insumo[i].get_text().strip()
                preco_insumo = precos[i].get_text().strip()
                # if not esgotado[i].get_text() == 'Esgotado':
                insumos_preco = {nome_insumo: preco_insumo}
                lista_insumos.append(insumos_preco)

            if len(proximo) > 0:
                num_pagina += 1
                url = 'http://loja.lamasbrewshop.com.br/insumos/malte-cereais.html?' + 'p=%d' % num_pagina
                requisicao, status = request_site(url, headers, num_pagina)
                # res = requests.get(url, headers=headers)
                # status_request = res.raise_for_status()
            else:
                break
        else:
            break

    return lista_insumos


def gera_planilha(lista_insumos, wb, planilha, site):
    """
    Funcao que exporta a lista de insumos para uma planilha no formato xlsx
    :param lista_insumos: a lista com os insumos depois do scraping
    :param wb: WORKBOOK, arquivo principal da planilha
    :param planilha: sheet (planilha) dentro do WORKBOOK
    :param site: nome do site que está sendo realizado o scraping
    :return: sem retorno. Grava a planilha.
    """
    for numeroLinha in range(2, len(lista_insumos)):
        insumo_preco = str(lista_insumos[numeroLinha]).split(":")

        planilha['A1'] = 'Insumo'
        planilha['B1'] = 'Preço'
        # insere na primeira coluna o nome dos insumos
        planilha.cell(row=numeroLinha, column=1).value = insumo_preco[0].replace('{', '').replace('\'', '')
        # insere na segunda coluna o preco dos insumos
        planilha.cell(row=numeroLinha, column=2).value = insumo_preco[1].replace('\'', '').replace('}', '')

    wb.save(site + '.xlsx')


num_pagina = 1
url = 'http://loja.lamasbrewshop.com.br/insumos/malte-cereais.html?' + 'p=%d' % num_pagina  # maltes


lista_insumos = []

requisicao, status = request_site(url, headers, num_pagina)
lista_precos_insumos = scraping_insumos(requisicao, status, num_pagina, lista_insumos)
gera_planilha(lista_insumos, wb, sheet, NOME_SITE)

# for i in range(len(lista_precos_insumos)):
#     print(lista_precos_insumos[i])

# gera arquivo xlsx
# for numeroLinha in range(2, len(lista_insumos)):
#     nome_valor = str(lista_insumos[numeroLinha]).split(":")
#
#     sheet.cell(row=numeroLinha, column=1).value = nome_valor[0].replace('{', '').replace('\'', '')
#     sheet.cell(row=numeroLinha, column=2).value = nome_valor[1].replace('\'', '').replace('}', '')
#
# wb.save('maltes_LAMAS.xlsx')






